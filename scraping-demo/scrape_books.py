# -*- coding: utf-8 -*-
"""
書籍情報スクレイピング＆Excelレポート自動生成デモ

対象サイト: https://books.toscrape.com/
（スクレイピング練習用に公開されているデモサイトです）

実行すると、書籍一覧を収集して「整形済みのExcelレポート」を出力します。
- データ一覧シート（フィルタ・ウィンドウ枠固定・書式設定済み）
- サマリーシート（件数・平均価格・評価別集計）

使い方:
    pip install requests beautifulsoup4 openpyxl
    python scrape_books.py
"""

import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://books.toscrape.com/catalogue/page-{}.html"
PAGES_TO_SCRAPE = 3          # 収集するページ数（1ページ20件）
REQUEST_INTERVAL_SEC = 1.0   # サーバーに負荷をかけないための待機時間

RATING_MAP = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (portfolio-demo; polite scraper)"
}


def scrape_page(page_num: int) -> list[dict]:
    """一覧ページ1枚分の書籍データを取得する"""
    url = BASE_URL.format(page_num)
    res = requests.get(url, headers=HEADERS, timeout=30)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "html.parser")

    books = []
    for article in soup.select("article.product_pod"):
        title = article.h3.a["title"]
        price_text = article.select_one(".price_color").get_text(strip=True)
        # 通貨記号を除去して数値化（例: "£51.77" -> 51.77）
        price = float(price_text.lstrip("£Â"))
        rating_class = article.select_one(".star-rating")["class"]
        rating = RATING_MAP.get(rating_class[1], 0)
        in_stock = "In stock" in article.select_one(".availability").get_text()
        detail_url = "https://books.toscrape.com/catalogue/" + article.h3.a["href"]

        books.append(
            {
                "タイトル": title,
                "価格(£)": price,
                "評価(星)": rating,
                "在庫": "あり" if in_stock else "なし",
                "詳細URL": detail_url,
            }
        )
    return books


def build_excel(books: list[dict], output_path: str) -> None:
    """収集データから整形済みExcelレポートを生成する"""
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- データ一覧シート ----
    ws = wb.active
    ws.title = "データ一覧"

    columns = list(books[0].keys())
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for book in books:
        ws.append([book[c] for c in columns])

    widths = [50, 10, 10, 8, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(columns)):
        for cell in row:
            cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = "0.00"          # 価格
        row[2].alignment = Alignment(horizontal="center")  # 評価
        row[3].alignment = Alignment(horizontal="center")  # 在庫

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    # ---- サマリーシート ----
    summary = wb.create_sheet("サマリー")
    prices = [b["価格(£)"] for b in books]

    summary_rows = [
        ("収集件数", len(books)),
        ("平均価格(£)", round(sum(prices) / len(prices), 2)),
        ("最高価格(£)", max(prices)),
        ("最低価格(£)", min(prices)),
    ]
    summary.append(("項目", "値"))
    for row in summary_rows:
        summary.append(row)

    summary.append(())
    summary.append(("評価(星)", "件数"))
    for star in range(5, 0, -1):
        count = sum(1 for b in books if b["評価(星)"] == star)
        summary.append((f"★{star}", count))

    for cell in list(summary[1]) + list(summary[6]):
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 12

    wb.save(output_path)


def main() -> None:
    all_books: list[dict] = []
    for page in range(1, PAGES_TO_SCRAPE + 1):
        print(f"ページ {page}/{PAGES_TO_SCRAPE} を取得中...")
        all_books.extend(scrape_page(page))
        time.sleep(REQUEST_INTERVAL_SEC)

    output = "books_report.xlsx"
    build_excel(all_books, output)
    print(f"完了: {len(all_books)} 件を収集し、{output} に出力しました。")


if __name__ == "__main__":
    main()
